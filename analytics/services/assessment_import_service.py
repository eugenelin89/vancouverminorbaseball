from __future__ import annotations

import hashlib
import json
import zipfile
from copy import deepcopy
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from django.conf import settings
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from analytics.models import (
    ASSESSMENT_CONFLICT_NONE,
    ASSESSMENT_CONFLICT_RESOLVED,
    ASSESSMENT_CONFLICT_UNRESOLVED,
    ASSESSMENT_IMPORT_ROW_AMBIGUOUS,
    ASSESSMENT_IMPORT_ROW_COMMITTED,
    ASSESSMENT_IMPORT_ROW_INVALID,
    ASSESSMENT_IMPORT_ROW_MATCHED,
    ASSESSMENT_IMPORT_ROW_SKIPPED,
    ASSESSMENT_IMPORT_ROW_UNMATCHED,
    ASSESSMENT_IMPORT_STATUS_COMMITTED,
    ASSESSMENT_IMPORT_STATUS_FAILED,
    ASSESSMENT_IMPORT_STATUS_PREVIEWED,
    ASSESSMENT_MATCH_AMBIGUOUS,
    ASSESSMENT_MATCH_MATCHED,
    ASSESSMENT_MATCH_UNMATCHED,
    ASSESSMENT_STATUS_COMMITTED,
    ASSESSMENT_STATUS_DRAFT,
    ASSESSMENT_VALIDATION_INVALID,
    ASSESSMENT_VALIDATION_VALID,
    ASSESSMENT_VALUE_SOURCE_IMPORTED,
    ASSESSMENT_VALUE_SOURCE_MANUAL_CORRECTED,
    ASSESSMENT_VALUE_TYPE_NUMBER,
    ASSESSMENT_VALUE_TYPE_RATING,
    AssessmentImportBatch,
    AssessmentImportRow,
    AssessmentImportTemplate,
    AssessmentMetricDefinition,
    AssessmentScoringProfile,
    AssessmentTemplate,
    AssessmentTemplateMetric,
    AssessmentValue,
    AssessmentValueCorrection,
    PlayerAssessment,
)
from analytics.services.assessment_matching_service import (
    MATCH_AMBIGUOUS,
    MATCH_UNMATCHED,
    match_player_for_assessment,
    normalize_assessment_name,
)
from players.models import Player

BOOTSTRAP_2026_13U_ASSESSMENT_KEY = "2026-13u-house-assessment"
BOOTSTRAP_2026_13U_IMPORT_KEY = "2026-13u-house-assessment-xlsx"

ZERO_ALLOW = "allow"
ZERO_TREAT_AS_MISSING = "treat_as_missing"
ZERO_WARNING = "warning"
ZERO_ERROR = "error"
ZERO_POLICIES = {ZERO_ALLOW, ZERO_TREAT_AS_MISSING, ZERO_WARNING, ZERO_ERROR}

BLANK_PRESERVE = "preserve_existing"
BLANK_CLEAR = "clear_existing_imported_value"
BLANK_IGNORE_CREATE = "ignore_on_create"
BLANK_REQUIRED_ERROR = "error_if_required"
BLANK_POLICIES = {
    BLANK_PRESERVE,
    BLANK_CLEAR,
    BLANK_IGNORE_CREATE,
    BLANK_REQUIRED_ERROR,
}

METRIC_ACTION_CREATE = "create"
METRIC_ACTION_UPDATE = "update"
METRIC_ACTION_UNCHANGED = "unchanged"
METRIC_ACTION_CLEAR = "clear"
METRIC_ACTION_SKIP = "skip"
METRIC_ACTION_PROTECTED_MANUAL = "protected_manual"
METRIC_ACTION_CONFLICT = "conflict"
METRIC_ACTION_INVALID = "invalid"

DEFAULT_MAX_UPLOAD_BYTES = 10 * 1024 * 1024
DEFAULT_MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
DEFAULT_MAX_WORKSHEETS = 12
DEFAULT_MAX_ROWS = 500
DEFAULT_MAX_COLUMNS = 50
DEFAULT_MAX_CELL_TEXT_LENGTH = 500


def _number_metric(
    header,
    key,
    category,
    *,
    direction="neutral",
    min_value=None,
    max_value=None,
    zero_policy=ZERO_ERROR,
):
    return {
        "header": header,
        "key": key,
        "category": category,
        "value_type": ASSESSMENT_VALUE_TYPE_NUMBER,
        "direction": direction,
        "required_header": True,
        "required_value": False,
        "min_value": min_value,
        "max_value": max_value,
        "unit": "",
        "unit_status": "unverified",
        "unit_source": "",
        "zero_policy": zero_policy,
        "blank_policy": BLANK_CLEAR,
    }


def _rating_metric(header, key, category):
    return {
        "header": header,
        "key": key,
        "category": category,
        "value_type": ASSESSMENT_VALUE_TYPE_RATING,
        "direction": "higher",
        "required_header": True,
        "required_value": False,
        "rating_scale_min": 1,
        "rating_scale_max": 3,
        "integer_only": True,
        "allowed_choices": [1, 2, 3],
        "unit": "",
        "unit_status": "not_applicable",
        "zero_policy": ZERO_ERROR,
        "blank_policy": BLANK_CLEAR,
    }


def _text_metric(header, key, category):
    return {
        "header": header,
        "key": key,
        "category": category,
        "value_type": "text",
        "required_header": True,
        "required_value": False,
        "unit": "",
        "unit_status": "not_applicable",
        "zero_policy": ZERO_ALLOW,
        "blank_policy": BLANK_CLEAR,
    }


DEFAULT_2026_13U_DATA_SHEETS = [
    {
        "name": "Assessment Data",
        "required": True,
        "header_row": 2,
        "identity_column": "Name",
        "category_row": 1,
        "max_rows": 500,
        "max_columns": 30,
        "metrics": [
            _number_metric(
                "Home to 1st",
                "home_to_1st",
                "Athleticism Evaluation",
                direction="lower",
                min_value=2,
                max_value=10,
            ),
            _number_metric(
                "Broad Jump",
                "broad_jump",
                "Athleticism Evaluation",
                direction="higher",
                min_value=1,
                max_value=200,
            ),
            _number_metric(
                "Lateral Jump",
                "lateral_jump",
                "Athleticism Evaluation",
                direction="higher",
                min_value=1,
                max_value=200,
            ),
            _number_metric(
                "Shotput",
                "shotput",
                "Athleticism Evaluation",
                direction="higher",
                min_value=1,
                max_value=1000,
            ),
            _number_metric(
                "Bat Speed",
                "bat_speed",
                "Hitting Objective Evaluation",
                direction="higher",
                min_value=1,
                max_value=150,
                zero_policy=ZERO_TREAT_AS_MISSING,
            ),
            _number_metric(
                "Time 2 Contact",
                "time_to_contact",
                "Hitting Objective Evaluation",
                direction="lower",
                min_value="0.01",
                max_value=2,
                zero_policy=ZERO_TREAT_AS_MISSING,
            ),
            _number_metric(
                "Exit Velocity Avg.",
                "exit_velocity_avg",
                "Hitting Objective Evaluation",
                direction="higher",
                min_value=1,
                max_value=150,
                zero_policy=ZERO_TREAT_AS_MISSING,
            ),
            _number_metric(
                "Exit Velocity Max",
                "exit_velocity_max",
                "Hitting Objective Evaluation",
                direction="higher",
                min_value=1,
                max_value=150,
                zero_policy=ZERO_TREAT_AS_MISSING,
            ),
            _rating_metric(
                "Athletic Stance", "athletic_stance", "Hitting Subjective Evaluation"
            ),
            _rating_metric(
                "Balance Stride", "balance_stride", "Hitting Subjective Evaluation"
            ),
            _rating_metric(
                "Barrel Level", "barrel_level", "Hitting Subjective Evaluation"
            ),
            _rating_metric(
                "Launch Position", "launch_position", "Hitting Subjective Evaluation"
            ),
            _rating_metric(
                "Follow Through", "follow_through", "Hitting Subjective Evaluation"
            ),
            _rating_metric(
                "Readiness", "fielding_readiness", "Fielding and Throwing Evaluation"
            ),
            _rating_metric(
                "Footwork", "fielding_footwork", "Fielding and Throwing Evaluation"
            ),
            _rating_metric(
                "Glovework", "fielding_glovework", "Fielding and Throwing Evaluation"
            ),
            _rating_metric(
                "Athleticism",
                "fielding_athleticism",
                "Fielding and Throwing Evaluation",
            ),
            _rating_metric(
                "Fundamental Throwing",
                "fundamental_throwing",
                "Fielding and Throwing Evaluation",
            ),
        ],
    },
    {
        "name": "Pitching Data",
        "required": False,
        "header_row": 2,
        "identity_column": "Name",
        "max_rows": 500,
        "max_columns": 20,
        "metrics": [
            _number_metric(
                "Velocity Avg.",
                "pitching_velocity_avg",
                "Pitching Data",
                direction="higher",
                min_value=1,
                max_value=120,
            ),
            _number_metric(
                "Velocity Max",
                "pitching_velocity_max",
                "Pitching Data",
                direction="higher",
                min_value=1,
                max_value=120,
            ),
            _text_metric("Pitch 1", "pitch_1", "Pitching Data"),
            _text_metric("Pitch 2", "pitch_2", "Pitching Data"),
            _text_metric("Pitch 3", "pitch_3", "Pitching Data"),
            _text_metric("Pitch 4", "pitch_4", "Pitching Data"),
            _rating_metric(
                "Athletic Movement", "pitching_athletic_movement", "Pitching Data"
            ),
            _rating_metric("Body Control", "pitching_body_control", "Pitching Data"),
            _rating_metric("Direction", "pitching_direction", "Pitching Data"),
            _rating_metric("Repeatability", "pitching_repeatability", "Pitching Data"),
            _rating_metric("Command2", "pitching_command", "Pitching Data"),
        ],
    },
]

DEFAULT_2026_13U_RANKING_SHEETS = ["Ranking", "Pitcher Ranking"]


@dataclass(frozen=True)
class AssessmentPreviewSummary:
    rows: int
    valid_player_rows: int
    matched: int
    unmatched: int
    ambiguous: int
    invalid: int
    skipped: int
    conflicts: int
    creates: int
    updates: int
    unchanged: int
    clears: int
    protected_manual: int
    workbook_errors: int
    workbook_warnings: int
    acknowledgement_required: bool
    acknowledgement_complete: bool
    checksum_seen_before: bool

    @property
    def structurally_ready(self) -> bool:
        return (
            self.valid_player_rows > 0
            and self.workbook_errors == 0
            and self.invalid == 0
            and self.unmatched == 0
            and self.ambiguous == 0
            and self.conflicts == 0
        )

    @property
    def can_commit(self) -> bool:
        return self.structurally_ready and (
            not self.acknowledgement_required or self.acknowledgement_complete
        )


@dataclass(frozen=True)
class AssessmentCommitResult:
    processed: int
    created: int
    updated: int
    unchanged: int
    skipped: int
    values_created: int
    values_updated: int
    values_cleared: int
    values_unchanged: int
    values_protected: int


def _issue(code, message, *, blocking=False, requires_ack=False, **context):
    issue = {
        "code": code,
        "message": message,
        "blocking": bool(blocking),
        "requires_ack": bool(requires_ack),
    }
    if context:
        issue["context"] = context
    return issue


def normalize_sheet_name(value: str) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def normalize_header(value: str) -> str:
    normalized = normalize_sheet_name(value)
    return "_".join(part for part in normalized.replace(".", " ").split() if part)


def _config_json(config: dict) -> str:
    return json.dumps(config, sort_keys=True, separators=(",", ":"), default=str)


def config_checksum(config: dict) -> str:
    return hashlib.sha256(_config_json(config).encode("utf-8")).hexdigest()


def _workbook_bytes(file_obj: BinaryIO) -> bytes:
    position = file_obj.tell() if hasattr(file_obj, "tell") else None
    content = file_obj.read()
    if position is not None and hasattr(file_obj, "seek"):
        file_obj.seek(position)
    return content


def workbook_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _limits(config: dict) -> dict:
    configured = config.get("limits", {})
    return {
        "max_upload_bytes": min(
            int(configured.get("max_upload_bytes", DEFAULT_MAX_UPLOAD_BYTES)),
            int(
                getattr(
                    settings,
                    "ANALYTICS_ASSESSMENT_MAX_UPLOAD_BYTES",
                    DEFAULT_MAX_UPLOAD_BYTES,
                )
            ),
        ),
        "max_worksheets": int(configured.get("max_worksheets", DEFAULT_MAX_WORKSHEETS)),
        "max_archive_uncompressed_bytes": min(
            int(
                configured.get(
                    "max_archive_uncompressed_bytes",
                    DEFAULT_MAX_UNCOMPRESSED_BYTES,
                )
            ),
            int(
                getattr(
                    settings,
                    "ANALYTICS_ASSESSMENT_MAX_UNCOMPRESSED_BYTES",
                    DEFAULT_MAX_UNCOMPRESSED_BYTES,
                )
            ),
        ),
        "max_rows": int(configured.get("max_rows", DEFAULT_MAX_ROWS)),
        "max_columns": int(configured.get("max_columns", DEFAULT_MAX_COLUMNS)),
        "max_cell_text_length": int(
            configured.get("max_cell_text_length", DEFAULT_MAX_CELL_TEXT_LENGTH)
        ),
    }


def _load_workbook_from_bytes(content: bytes, config: dict):
    limits = _limits(config)
    if len(content) > limits["max_upload_bytes"]:
        raise ValidationError("Workbook exceeds the configured upload size limit.")
    if not zipfile.is_zipfile(BytesIO(content)):
        raise ValidationError("Workbook is not a valid .xlsx file.")
    with zipfile.ZipFile(BytesIO(content)) as archive:
        members = archive.infolist()
        names = {member.filename for member in members}
        if (
            sum(member.file_size for member in members)
            > limits["max_archive_uncompressed_bytes"]
        ):
            raise ValidationError(
                "Workbook expands beyond the configured safe processing limit."
            )
        if any(name.lower().endswith("vbaproject.bin") for name in names):
            raise ValidationError("Macro-enabled workbooks are not supported.")
        if any(name.startswith("xl/externalLinks/") for name in names):
            raise ValidationError("Workbooks with external links are not supported.")
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ValidationError(
            "Workbook could not be read as a valid .xlsx file."
        ) from exc
    if len(workbook.sheetnames) > limits["max_worksheets"]:
        workbook.close()
        raise ValidationError("Workbook contains too many worksheets.")
    return workbook


def _worksheet_by_name(workbook, configured_name: str):
    normalized = normalize_sheet_name(configured_name)
    for sheet_name in workbook.sheetnames:
        if normalize_sheet_name(sheet_name) == normalized:
            return workbook[sheet_name]
    return None


def _header_map(row_values: list) -> tuple[dict[str, int], list[str]]:
    mapping = {}
    duplicates = []
    for index, value in enumerate(row_values):
        if value in (None, ""):
            continue
        normalized = normalize_header(value)
        if normalized in mapping:
            duplicates.append(str(value).strip())
        else:
            mapping[normalized] = index
    return mapping, duplicates


def _header_candidates(config: dict) -> list[str]:
    return [config.get("header", ""), *config.get("header_aliases", [])]


def _find_header_index(headers: dict[str, int], config: dict) -> int | None:
    for candidate in _header_candidates(config):
        index = headers.get(normalize_header(candidate))
        if index is not None:
            return index
    return None


def _decimal_or_none(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        result = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if not result.is_finite():
        return None
    return result


def _snapshot_value(
    metric_config: dict,
    raw_value,
    *,
    sheet_name: str,
    row_number: int,
    column_index: int,
    max_cell_text_length: int,
) -> dict:
    value_type = metric_config.get("value_type") or ASSESSMENT_VALUE_TYPE_NUMBER
    raw_text = "" if raw_value is None else str(raw_value).strip()
    snapshot = {
        "metric_key": metric_config["key"],
        "header": metric_config["header"],
        "value_type": value_type,
        "unit": metric_config.get("unit", ""),
        "unit_status": metric_config.get("unit_status", "not_applicable"),
        "unit_source": metric_config.get("unit_source", ""),
        "rating_scale_min": metric_config.get("rating_scale_min"),
        "rating_scale_max": metric_config.get("rating_scale_max"),
        "zero_policy": metric_config.get("zero_policy", ZERO_ALLOW),
        "blank_policy": metric_config.get("blank_policy", BLANK_PRESERVE),
        "raw_value": raw_text,
        "normalized_value": "",
        "is_blank": raw_value in (None, ""),
        "source_sheet": sheet_name,
        "source_row": row_number,
        "source_column": get_column_letter(column_index + 1),
        "source_header": metric_config["header"],
        "errors": [],
        "warnings": [],
        "transformations": [],
    }
    if len(raw_text) > max_cell_text_length:
        snapshot["errors"].append(
            _issue(
                "cell_text_too_long",
                f"{metric_config['header']} exceeds the cell text limit.",
                blocking=True,
            )
        )
        return snapshot
    if snapshot["is_blank"]:
        if (
            metric_config.get("required_value")
            or snapshot["blank_policy"] == BLANK_REQUIRED_ERROR
        ):
            snapshot["errors"].append(
                _issue(
                    "required_value_missing",
                    f"{metric_config['header']} is required for this row.",
                    blocking=True,
                )
            )
        return snapshot
    if isinstance(raw_value, str) and raw_value.startswith("="):
        snapshot["errors"].append(
            _issue(
                "formula_not_supported",
                f"{metric_config['header']} contains a formula instead of a value.",
                blocking=True,
            )
        )
        return snapshot
    if value_type not in {ASSESSMENT_VALUE_TYPE_NUMBER, ASSESSMENT_VALUE_TYPE_RATING}:
        snapshot["text_value"] = raw_text
        snapshot["normalized_value"] = raw_text
        return snapshot

    decimal_value = _decimal_or_none(raw_value)
    if decimal_value is None:
        snapshot["errors"].append(
            _issue(
                "invalid_numeric_value",
                f"{metric_config['header']} is not a valid finite number.",
                blocking=True,
            )
        )
        return snapshot

    zero_policy = snapshot["zero_policy"]
    if zero_policy not in ZERO_POLICIES:
        snapshot["errors"].append(
            _issue(
                "invalid_zero_policy",
                f"{metric_config['header']} has an invalid zero policy.",
                blocking=True,
            )
        )
        return snapshot
    if decimal_value == 0:
        if zero_policy == ZERO_ERROR:
            snapshot["errors"].append(
                _issue(
                    "zero_not_allowed",
                    f"{metric_config['header']} cannot be zero.",
                    blocking=True,
                )
            )
            return snapshot
        if zero_policy == ZERO_TREAT_AS_MISSING:
            snapshot["is_blank"] = True
            snapshot["warnings"].append(
                _issue(
                    "zero_treated_as_missing",
                    f"{metric_config['header']} zero will be treated as missing.",
                    requires_ack=True,
                )
            )
            snapshot["transformations"].append(
                {
                    "kind": "zero_to_missing",
                    "reason": "Configured zero policy treats implausible zero as missing.",
                    "policy": ZERO_TREAT_AS_MISSING,
                }
            )
            return snapshot
        if zero_policy == ZERO_WARNING:
            snapshot["warnings"].append(
                _issue(
                    "zero_requires_review",
                    f"{metric_config['header']} contains zero and requires review.",
                    requires_ack=True,
                )
            )

    min_value = _decimal_or_none(metric_config.get("min_value"))
    max_value = _decimal_or_none(metric_config.get("max_value"))
    if min_value is not None and decimal_value < min_value:
        snapshot["errors"].append(
            _issue(
                "value_below_minimum",
                f"{metric_config['header']} is below the configured minimum {min_value}.",
                blocking=True,
            )
        )
    if max_value is not None and decimal_value > max_value:
        snapshot["errors"].append(
            _issue(
                "value_above_maximum",
                f"{metric_config['header']} is above the configured maximum {max_value}.",
                blocking=True,
            )
        )
    if (
        metric_config.get("integer_only")
        and decimal_value != decimal_value.to_integral_value()
    ):
        snapshot["errors"].append(
            _issue(
                "integer_required",
                f"{metric_config['header']} must be a whole number.",
                blocking=True,
            )
        )
    allowed_choices = {
        Decimal(str(value)) for value in metric_config.get("allowed_choices", [])
    }
    if allowed_choices and decimal_value not in allowed_choices:
        snapshot["errors"].append(
            _issue(
                "value_not_allowed",
                f"{metric_config['header']} must be one of {sorted(allowed_choices)}.",
                blocking=True,
            )
        )
    if value_type == ASSESSMENT_VALUE_TYPE_RATING:
        rating_min = _decimal_or_none(metric_config.get("rating_scale_min"))
        rating_max = _decimal_or_none(metric_config.get("rating_scale_max"))
        if rating_min is None or rating_max is None:
            snapshot["errors"].append(
                _issue(
                    "rating_scale_missing",
                    f"{metric_config['header']} has no configured rating scale.",
                    blocking=True,
                )
            )
        elif decimal_value < rating_min or decimal_value > rating_max:
            snapshot["errors"].append(
                _issue(
                    "rating_out_of_range",
                    f"{metric_config['header']} must be between {rating_min} and {rating_max}.",
                    blocking=True,
                )
            )
    if snapshot["unit_status"] == "unverified":
        snapshot["warnings"].append(
            _issue(
                "unit_unverified",
                f"{metric_config['header']} unit is not confirmed.",
                requires_ack=True,
            )
        )
    if not snapshot["errors"]:
        snapshot["numeric_value"] = str(decimal_value)
        snapshot["normalized_value"] = str(decimal_value)
    return snapshot


def _configured_header_names(sheet_config: dict) -> set[str]:
    names = {normalize_header(sheet_config.get("identity_column", "Name"))}
    for alias in sheet_config.get("identity_aliases", []):
        names.add(normalize_header(alias))
    for metric in sheet_config.get("metrics", []):
        names.update(normalize_header(value) for value in _header_candidates(metric))
    for identifier in sheet_config.get("source_identifiers", []):
        names.update(
            normalize_header(value) for value in _header_candidates(identifier)
        )
    return names


def _parse_sheet(
    workbook, sheet_config: dict, config: dict
) -> tuple[list[dict], list, list]:
    worksheet = _worksheet_by_name(workbook, sheet_config["name"])
    errors = []
    warnings = []
    if worksheet is None:
        issue = _issue(
            (
                "required_sheet_missing"
                if sheet_config.get("required", True)
                else "optional_sheet_missing"
            ),
            f"{'Required' if sheet_config.get('required', True) else 'Optional'} worksheet is missing: {sheet_config['name']}.",
            blocking=sheet_config.get("required", True),
            requires_ack=not sheet_config.get("required", True),
            sheet=sheet_config["name"],
        )
        (errors if issue["blocking"] else warnings).append(issue)
        return [], errors, warnings

    limits = _limits(config)
    max_rows = min(
        int(sheet_config.get("max_rows", limits["max_rows"])), limits["max_rows"]
    )
    max_columns = min(
        int(sheet_config.get("max_columns", limits["max_columns"])),
        limits["max_columns"],
    )
    if worksheet.max_row > max_rows:
        errors.append(
            _issue(
                "worksheet_row_limit",
                f"Worksheet {worksheet.title} exceeds the {max_rows}-row limit.",
                blocking=True,
                sheet=worksheet.title,
            )
        )
        return [], errors, warnings
    if worksheet.max_column > max_columns:
        errors.append(
            _issue(
                "worksheet_column_limit",
                f"Worksheet {worksheet.title} exceeds the {max_columns}-column limit.",
                blocking=True,
                sheet=worksheet.title,
            )
        )
        return [], errors, warnings

    header_row = int(sheet_config.get("header_row", 1))
    if header_row < 1 or header_row > worksheet.max_row:
        errors.append(
            _issue(
                "header_row_missing",
                f"Required header row {header_row} is missing from {worksheet.title}.",
                blocking=True,
                sheet=worksheet.title,
            )
        )
        return [], errors, warnings
    header_cells = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            max_col=worksheet.max_column,
            values_only=True,
        ),
        (),
    )
    if not any(value not in (None, "") for value in header_cells):
        errors.append(
            _issue(
                "header_row_empty",
                f"Required header row is empty in {worksheet.title}.",
                blocking=True,
                sheet=worksheet.title,
            )
        )
        return [], errors, warnings
    headers, duplicate_headers = _header_map(list(header_cells))
    for duplicate_header in duplicate_headers:
        errors.append(
            _issue(
                "duplicate_header",
                f"Duplicate header {duplicate_header} in {worksheet.title}.",
                blocking=True,
                sheet=worksheet.title,
            )
        )

    identity_config = {
        "header": sheet_config.get("identity_column", "Name"),
        "header_aliases": sheet_config.get("identity_aliases", []),
    }
    identity_index = _find_header_index(headers, identity_config)
    if identity_index is None:
        errors.append(
            _issue(
                "identity_header_missing",
                f"Identity header {identity_config['header']} is missing from {worksheet.title}.",
                blocking=True,
                sheet=worksheet.title,
            )
        )

    metric_indexes = []
    for metric_config in sheet_config.get("metrics", []):
        metric_index = _find_header_index(headers, metric_config)
        if metric_index is None:
            issue = _issue(
                (
                    "required_metric_header_missing"
                    if metric_config.get("required_header", True)
                    else "optional_metric_header_missing"
                ),
                f"Expected header {metric_config['header']} is missing from {worksheet.title}.",
                blocking=metric_config.get("required_header", True),
                requires_ack=not metric_config.get("required_header", True),
                sheet=worksheet.title,
                metric=metric_config["key"],
            )
            (errors if issue["blocking"] else warnings).append(issue)
        else:
            metric_indexes.append((metric_config, metric_index))

    identifier_indexes = []
    for identifier_config in sheet_config.get("source_identifiers", []):
        identifier_index = _find_header_index(headers, identifier_config)
        if identifier_index is not None:
            identifier_indexes.append((identifier_config, identifier_index))
        elif identifier_config.get("required_header"):
            errors.append(
                _issue(
                    "source_identifier_header_missing",
                    f"Source identifier header {identifier_config['header']} is missing from {worksheet.title}.",
                    blocking=True,
                )
            )

    expected_headers = _configured_header_names(sheet_config)
    for normalized, index in headers.items():
        if normalized not in expected_headers:
            warnings.append(
                _issue(
                    "unexpected_column",
                    f"Unexpected column {header_cells[index]} in {worksheet.title} will be ignored.",
                    requires_ack=True,
                    sheet=worksheet.title,
                    column=get_column_letter(index + 1),
                )
            )

    if identity_index is None or errors:
        return [], errors, warnings

    parsed_rows = []
    seen_identities: dict[str, list[dict]] = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        raw_identity = row[identity_index] if identity_index < len(row) else None
        has_other_data = any(value not in (None, "") for value in row)
        if raw_identity in (None, "") and not has_other_data:
            continue
        normalized_identity = normalize_assessment_name(raw_identity)
        row_errors = []
        if not normalized_identity:
            row_errors.append(
                _issue(
                    "identity_value_missing",
                    f"Row {row_number} in {worksheet.title} has data but no player identity.",
                    blocking=True,
                )
            )
            normalized_identity = (
                f"__missing__:{normalize_sheet_name(worksheet.title)}:{row_number}"
            )
        values = []
        row_warnings = []
        for metric_config, metric_index in metric_indexes:
            raw_value = row[metric_index] if metric_index < len(row) else None
            snapshot = _snapshot_value(
                metric_config,
                raw_value,
                sheet_name=worksheet.title,
                row_number=row_number,
                column_index=metric_index,
                max_cell_text_length=limits["max_cell_text_length"],
            )
            row_errors.extend(snapshot["errors"])
            row_warnings.extend(snapshot["warnings"])
            values.append(snapshot)
        source_identifiers = []
        for identifier_config, identifier_index in identifier_indexes:
            raw_value = row[identifier_index] if identifier_index < len(row) else None
            if raw_value in (None, ""):
                continue
            source_identifiers.append(
                {
                    "source": identifier_config.get("source", ""),
                    "identifier_type": identifier_config.get("identifier_type", ""),
                    "identifier_value": str(raw_value).strip(),
                }
            )
        parsed_row = {
            "normalized_identity": normalized_identity,
            "raw_identity": "" if raw_identity is None else str(raw_identity).strip(),
            "source_rows": [{"sheet": worksheet.title, "row": row_number}],
            "source_identifiers": source_identifiers,
            "values": values,
            "errors": row_errors,
            "warnings": row_warnings,
        }
        parsed_rows.append(parsed_row)
        seen_identities.setdefault(normalized_identity, []).append(parsed_row)

    for normalized_identity, duplicate_rows in seen_identities.items():
        if len(duplicate_rows) < 2:
            continue
        for duplicate_row in duplicate_rows:
            duplicate_row["errors"].append(
                _issue(
                    "duplicate_identity_in_sheet",
                    f"Player identity appears more than once in {worksheet.title}.",
                    blocking=True,
                    normalized_identity=normalized_identity,
                )
            )
    return parsed_rows, errors, warnings


def _combine_component_rows(component_rows: list[dict]) -> list[dict]:
    combined: dict[str, dict] = {}
    slug_identities: dict[str, set[str]] = {}
    source_identifier_rows: dict[tuple[str, str, str], list[dict]] = {}
    for component in component_rows:
        normalized_identity = component["normalized_identity"]
        row = combined.setdefault(
            normalized_identity,
            {
                "row_key": hashlib.sha256(
                    normalized_identity.encode("utf-8")
                ).hexdigest()[:32],
                "normalized_identity": normalized_identity,
                "raw_identity": component["raw_identity"],
                "source_rows": [],
                "source_identifiers": [],
                "values": [],
                "errors": [],
                "warnings": [],
            },
        )
        row["source_rows"].extend(component["source_rows"])
        row["source_identifiers"].extend(component["source_identifiers"])
        row["errors"].extend(component["errors"])
        row["warnings"].extend(component["warnings"])
        existing_values = {value["metric_key"]: value for value in row["values"]}
        for value in component["values"]:
            existing = existing_values.get(value["metric_key"])
            if existing and existing.get("raw_value") != value.get("raw_value"):
                row["errors"].append(
                    _issue(
                        "conflicting_duplicate_metric",
                        f"Conflicting source values exist for {value['header']}.",
                        blocking=True,
                    )
                )
            elif not existing:
                row["values"].append(value)
                existing_values[value["metric_key"]] = value
        slug_identities.setdefault(slugify(normalized_identity), set()).add(
            normalized_identity
        )
        for identifier in component["source_identifiers"]:
            key = (
                normalize_sheet_name(identifier.get("source", "")),
                normalize_sheet_name(identifier.get("identifier_type", "")),
                normalize_sheet_name(identifier.get("identifier_value", "")),
            )
            source_identifier_rows.setdefault(key, []).append(row)

    for identities in slug_identities.values():
        if len(identities) < 2:
            continue
        for identity in identities:
            combined[identity]["errors"].append(
                _issue(
                    "identity_slug_collision",
                    "Distinct player identities would collide under slug matching; manual correction is required.",
                    blocking=True,
                )
            )
    for identifier, rows in source_identifier_rows.items():
        unique_rows = {row["row_key"]: row for row in rows}
        if identifier == ("", "", "") or len(unique_rows) < 2:
            continue
        for row in unique_rows.values():
            row["errors"].append(
                _issue(
                    "duplicate_source_identifier",
                    "A source identifier is assigned to multiple workbook rows.",
                    blocking=True,
                )
            )
    return list(combined.values())


def parse_assessment_workbook(content: bytes, config_or_template) -> dict:
    """Parse a workbook using a frozen configuration without writing player data."""
    config = (
        config_or_template.config
        if isinstance(config_or_template, AssessmentImportTemplate)
        else deepcopy(config_or_template)
    )
    workbook = _load_workbook_from_bytes(content, config)
    workbook_errors = []
    workbook_warnings = []
    component_rows = []
    try:
        for sheet_config in config.get("sheets", []):
            rows, errors, warnings = _parse_sheet(workbook, sheet_config, config)
            component_rows.extend(rows)
            workbook_errors.extend(errors)
            workbook_warnings.extend(warnings)
    finally:
        workbook.close()
    rows = _combine_component_rows(component_rows)
    if not rows:
        workbook_errors.append(
            _issue(
                "no_player_rows",
                "No valid source player rows were parsed from the workbook.",
                blocking=True,
            )
        )
    return {
        "rows": rows,
        "errors": workbook_errors,
        "warnings": workbook_warnings,
        "ranking_sheets": config.get("ranking_sheets", []),
    }


def _value_snapshot(value: AssessmentValue | None) -> dict | None:
    if value is None:
        return None
    return {
        "id": value.pk,
        "numeric_value": (
            str(value.numeric_value) if value.numeric_value is not None else None
        ),
        "rating_value": (
            str(value.rating_value) if value.rating_value is not None else None
        ),
        "rating_scale_min": (
            str(value.rating_scale_min) if value.rating_scale_min is not None else None
        ),
        "rating_scale_max": (
            str(value.rating_scale_max) if value.rating_scale_max is not None else None
        ),
        "text_value": value.text_value,
        "choice_value": value.choice_value,
        "raw_value": value.raw_value,
        "normalized_value": value.normalized_value,
        "unit": value.unit,
        "source_kind": value.source_kind,
        "is_manual_override": value.is_manual_override,
        "source_sheet": value.source_sheet,
        "source_header": value.source_header,
    }


def _existing_normalized(value: AssessmentValue) -> str:
    if value.rating_value is not None:
        return str(value.rating_value)
    if value.numeric_value is not None:
        return str(value.numeric_value)
    if value.choice_value:
        return value.choice_value
    return value.text_value


def _values_equal(value: AssessmentValue, snapshot: dict) -> bool:
    incoming = snapshot.get("normalized_value", "")
    if snapshot.get("value_type") in {
        ASSESSMENT_VALUE_TYPE_NUMBER,
        ASSESSMENT_VALUE_TYPE_RATING,
    }:
        existing_decimal = _decimal_or_none(_existing_normalized(value))
        incoming_decimal = _decimal_or_none(incoming)
        return existing_decimal is not None and existing_decimal == incoming_decimal
    return _existing_normalized(value) == incoming


def _plan_metric_changes(
    parsed_row: dict, *, event, player: Player | None
) -> list[dict]:
    existing_assessment = None
    existing_values = {}
    if player:
        existing_assessment = PlayerAssessment.objects.filter(
            player=player, event=event
        ).first()
        if existing_assessment:
            existing_values = {
                value.template_metric.metric.key: value
                for value in existing_assessment.values.select_related(
                    "template_metric__metric"
                )
            }
    changes = []
    for snapshot in parsed_row.get("values", []):
        existing = existing_values.get(snapshot["metric_key"])
        change = {
            "metric_key": snapshot["metric_key"],
            "header": snapshot["header"],
            "value_type": snapshot["value_type"],
            "unit": snapshot.get("unit", ""),
            "unit_status": snapshot.get("unit_status", "not_applicable"),
            "rating_scale_min": snapshot.get("rating_scale_min"),
            "rating_scale_max": snapshot.get("rating_scale_max"),
            "old_value": _value_snapshot(existing),
            "incoming_raw_value": snapshot.get("raw_value", ""),
            "incoming_normalized_value": snapshot.get("normalized_value", ""),
            "source_sheet": snapshot.get("source_sheet", ""),
            "source_row": snapshot.get("source_row"),
            "source_column": snapshot.get("source_column", ""),
            "source_header": snapshot.get("source_header", ""),
            "warnings": snapshot.get("warnings", []),
            "errors": snapshot.get("errors", []),
            "transformations": snapshot.get("transformations", []),
            "resolution": "",
        }
        if snapshot.get("errors"):
            change["action"] = METRIC_ACTION_INVALID
        elif snapshot.get("is_blank"):
            if existing and existing.is_manual_override:
                change["action"] = METRIC_ACTION_CONFLICT
                change["conflict"] = "manual_override_blank"
            elif existing and snapshot.get("blank_policy") == BLANK_CLEAR:
                change["action"] = METRIC_ACTION_CLEAR
            elif existing:
                change["action"] = METRIC_ACTION_UNCHANGED
            else:
                change["action"] = METRIC_ACTION_SKIP
        elif existing and existing.is_manual_override:
            if _values_equal(existing, snapshot):
                change["action"] = METRIC_ACTION_PROTECTED_MANUAL
            else:
                change["action"] = METRIC_ACTION_CONFLICT
                change["conflict"] = "manual_override_difference"
        elif existing is None:
            change["action"] = METRIC_ACTION_CREATE
        elif (
            _values_equal(existing, snapshot)
            and existing.unit == snapshot.get("unit", "")
            and (
                snapshot["value_type"] != ASSESSMENT_VALUE_TYPE_RATING
                or (
                    existing.rating_scale_min
                    == _decimal_or_none(snapshot.get("rating_scale_min"))
                    and existing.rating_scale_max
                    == _decimal_or_none(snapshot.get("rating_scale_max"))
                )
            )
        ):
            change["action"] = METRIC_ACTION_UNCHANGED
        else:
            change["action"] = METRIC_ACTION_UPDATE
        changes.append(change)
    return changes


def _match_status(match) -> str:
    if match.status == MATCH_AMBIGUOUS:
        return ASSESSMENT_MATCH_AMBIGUOUS
    if match.status == MATCH_UNMATCHED:
        return ASSESSMENT_MATCH_UNMATCHED
    return ASSESSMENT_MATCH_MATCHED


def _legacy_row_status(row) -> str:
    if row.status in {ASSESSMENT_IMPORT_ROW_SKIPPED, ASSESSMENT_IMPORT_ROW_COMMITTED}:
        return row.status
    if row.validation_status == ASSESSMENT_VALIDATION_INVALID:
        return ASSESSMENT_IMPORT_ROW_INVALID
    if row.conflict_status == ASSESSMENT_CONFLICT_UNRESOLVED:
        return ASSESSMENT_IMPORT_ROW_INVALID
    if row.match_status == ASSESSMENT_MATCH_AMBIGUOUS:
        return ASSESSMENT_IMPORT_ROW_AMBIGUOUS
    if row.match_status == ASSESSMENT_MATCH_UNMATCHED:
        return ASSESSMENT_IMPORT_ROW_UNMATCHED
    return ASSESSMENT_IMPORT_ROW_MATCHED


def _planned_row_action(row) -> str:
    if row.status == ASSESSMENT_IMPORT_ROW_SKIPPED:
        return "skip"
    actions = {change.get("action") for change in row.metric_changes}
    if METRIC_ACTION_INVALID in actions:
        return "invalid"
    if METRIC_ACTION_CONFLICT in actions:
        return "conflict"
    if row.match_status != ASSESSMENT_MATCH_MATCHED:
        return "needs_identity_resolution"
    if METRIC_ACTION_CREATE in actions and row.player_id:
        return "create"
    if actions & {METRIC_ACTION_UPDATE, METRIC_ACTION_CLEAR}:
        return "update"
    if actions <= {
        METRIC_ACTION_UNCHANGED,
        METRIC_ACTION_SKIP,
        METRIC_ACTION_PROTECTED_MANUAL,
    }:
        return "unchanged"
    return "skip"


def _summary_counts(batch: AssessmentImportBatch) -> AssessmentPreviewSummary:
    rows = list(batch.rows.all())
    metric_actions = [
        change.get("action") for row in rows for change in row.metric_changes
    ]
    valid_rows = [
        row
        for row in rows
        if row.status != ASSESSMENT_IMPORT_ROW_SKIPPED
        and row.validation_status == ASSESSMENT_VALIDATION_VALID
    ]
    acknowledgement_required = bool(batch.required_warning_codes)
    acknowledgement_complete = bool(
        not acknowledgement_required
        or (
            batch.warnings_acknowledged_at
            and batch.metadata.get("acknowledged_token") == batch.acknowledgement_token
        )
    )
    return AssessmentPreviewSummary(
        rows=len(rows),
        valid_player_rows=len(valid_rows),
        matched=sum(
            1 for row in valid_rows if row.match_status == ASSESSMENT_MATCH_MATCHED
        ),
        unmatched=sum(
            1 for row in valid_rows if row.match_status == ASSESSMENT_MATCH_UNMATCHED
        ),
        ambiguous=sum(
            1 for row in valid_rows if row.match_status == ASSESSMENT_MATCH_AMBIGUOUS
        ),
        invalid=sum(
            1
            for row in rows
            if row.status != ASSESSMENT_IMPORT_ROW_SKIPPED
            and row.validation_status == ASSESSMENT_VALIDATION_INVALID
        ),
        skipped=sum(1 for row in rows if row.status == ASSESSMENT_IMPORT_ROW_SKIPPED),
        conflicts=sum(
            1
            for row in rows
            if row.status != ASSESSMENT_IMPORT_ROW_SKIPPED
            and row.conflict_status == ASSESSMENT_CONFLICT_UNRESOLVED
        ),
        creates=metric_actions.count(METRIC_ACTION_CREATE),
        updates=metric_actions.count(METRIC_ACTION_UPDATE),
        unchanged=metric_actions.count(METRIC_ACTION_UNCHANGED),
        clears=metric_actions.count(METRIC_ACTION_CLEAR),
        protected_manual=metric_actions.count(METRIC_ACTION_PROTECTED_MANUAL),
        workbook_errors=len(batch.validation_errors),
        workbook_warnings=len(batch.validation_warnings),
        acknowledgement_required=acknowledgement_required,
        acknowledgement_complete=acknowledgement_complete,
        checksum_seen_before=bool(batch.preview_snapshot.get("checksum_seen_before")),
    )


def _warning_codes(batch: AssessmentImportBatch) -> list[str]:
    issues = list(batch.validation_warnings)
    for row in batch.rows.all():
        issues.extend(row.warnings)
        for change in row.metric_changes:
            issues.extend(change.get("warnings", []))
    return sorted(
        {
            issue.get("code")
            for issue in issues
            if issue.get("requires_ack") and issue.get("code")
        }
    )


def _acknowledgement_token(batch: AssessmentImportBatch) -> str:
    payload = {
        "batch": batch.pk,
        "workbook": batch.workbook_sha256,
        "config": batch.config_checksum,
        "preview_version": batch.preview_version,
        "warning_codes": batch.required_warning_codes,
        "row_state": list(
            batch.rows.order_by("pk").values(
                "pk",
                "player_id",
                "status",
                "match_status",
                "validation_status",
                "conflict_status",
                "metric_changes",
            )
        ),
    }
    return hashlib.sha256(_config_json(payload).encode("utf-8")).hexdigest()


def _refresh_batch_state(batch: AssessmentImportBatch, *, bump_version=True):
    if bump_version:
        batch.preview_version += 1
    batch.required_warning_codes = _warning_codes(batch)
    batch.acknowledgement_token = _acknowledgement_token(batch)
    batch.warnings_acknowledged_at = None
    batch.warnings_acknowledged_by = None
    batch.metadata.pop("acknowledged_token", None)
    summary = _summary_counts(batch)
    batch.import_summary = asdict(summary)
    batch.preview_snapshot = {
        **batch.preview_snapshot,
        "summary": asdict(summary),
        "mapping_version": batch.import_template.version,
        "config_checksum": batch.config_checksum,
    }
    batch.save(
        update_fields=[
            "preview_version",
            "required_warning_codes",
            "acknowledgement_token",
            "warnings_acknowledged_at",
            "warnings_acknowledged_by",
            "metadata",
            "import_summary",
            "preview_snapshot",
            "updated_at",
        ]
    )
    return summary


def summarize_import_batch(batch: AssessmentImportBatch) -> AssessmentPreviewSummary:
    """Return the current authoritative preview summary."""
    return _summary_counts(batch)


def _validate_template_compatibility(event, import_template):
    if not import_template.assessment_template_id:
        raise ValidationError(
            "Import template does not declare a compatible assessment template."
        )
    if import_template.assessment_template_id != event.template_id:
        raise ValidationError(
            "Import template is not compatible with the selected assessment event."
        )
    if event.scoring_profile_id:
        if event.scoring_profile.assessment_template_id != event.template_id:
            raise ValidationError(
                "Assessment event scoring profile is not compatible with its template."
            )


def create_assessment_import_batch(
    *, file_obj, event, import_template, uploaded_by
) -> AssessmentImportBatch:
    """Create an auditable preview batch without committing assessment values."""
    filename = Path(file_obj.name).name
    if not filename.lower().endswith(".xlsx"):
        raise ValidationError("Upload an .xlsx workbook.")
    _validate_template_compatibility(event, import_template)
    frozen_config = deepcopy(import_template.config)
    content = _workbook_bytes(file_obj)
    max_upload = _limits(frozen_config)["max_upload_bytes"]
    if len(content) > max_upload:
        raise ValidationError(
            f"Workbook exceeds the configured {max_upload}-byte upload limit."
        )
    checksum = workbook_sha256(content)
    checksum_seen_before = AssessmentImportBatch.objects.filter(
        event=event,
        workbook_sha256=checksum,
        status=ASSESSMENT_IMPORT_STATUS_COMMITTED,
    ).exists()
    batch = AssessmentImportBatch.objects.create(
        event=event,
        import_template=import_template,
        uploaded_by=uploaded_by,
        original_filename=filename,
        workbook_sha256=checksum,
        config_snapshot=frozen_config,
        config_checksum=config_checksum(frozen_config),
        preview_snapshot={"checksum_seen_before": checksum_seen_before},
    )
    try:
        parsed = parse_assessment_workbook(content, frozen_config)
        build_assessment_import_preview(batch=batch, parsed=parsed)
    except ValidationError:
        batch.status = ASSESSMENT_IMPORT_STATUS_FAILED
        batch.validation_errors = [
            _issue(
                "workbook_parse_failed",
                "Workbook could not be read safely. Verify the file and try again.",
                blocking=True,
            )
        ]
        batch.import_summary = {"errors": len(batch.validation_errors)}
        batch.save(
            update_fields=[
                "status",
                "validation_errors",
                "import_summary",
                "updated_at",
            ]
        )
    except Exception as exc:
        batch.status = ASSESSMENT_IMPORT_STATUS_FAILED
        batch.validation_errors = [
            _issue(
                "workbook_parse_failed",
                "Workbook parsing failed without creating assessment data.",
                blocking=True,
            )
        ]
        batch.metadata = {
            **batch.metadata,
            "failure_type": exc.__class__.__name__,
        }
        batch.import_summary = {"errors": len(batch.validation_errors)}
        batch.save(
            update_fields=[
                "status",
                "validation_errors",
                "metadata",
                "import_summary",
                "updated_at",
            ]
        )
    batch.refresh_from_db()
    return batch


@transaction.atomic
def build_assessment_import_preview(
    *, batch: AssessmentImportBatch, parsed: dict
) -> AssessmentPreviewSummary:
    """Persist structural validation, identity matches, and per-metric actions."""
    batch = AssessmentImportBatch.objects.select_for_update().get(pk=batch.pk)
    if batch.status == ASSESSMENT_IMPORT_STATUS_COMMITTED:
        raise ValidationError("Committed assessment imports cannot be previewed again.")
    batch.rows.all().delete()
    batch.validation_errors = list(parsed.get("errors", []))
    batch.validation_warnings = list(parsed.get("warnings", []))
    if batch.preview_snapshot.get("checksum_seen_before"):
        batch.validation_warnings.append(
            _issue(
                "duplicate_workbook_checksum",
                "This workbook checksum has already been committed for the event.",
                requires_ack=True,
            )
        )
    batch.status = ASSESSMENT_IMPORT_STATUS_PREVIEWED
    batch.save(
        update_fields=[
            "validation_errors",
            "validation_warnings",
            "status",
            "updated_at",
        ]
    )
    for parsed_row in parsed.get("rows", []):
        match = match_player_for_assessment(
            raw_name=parsed_row["raw_identity"],
            event=batch.event,
            source_identifiers=parsed_row.get("source_identifiers", []),
        )
        row_errors = list(parsed_row.get("errors", []))
        row_warnings = list(parsed_row.get("warnings", []))
        validation_status = (
            ASSESSMENT_VALIDATION_INVALID if row_errors else ASSESSMENT_VALIDATION_VALID
        )
        match_status = _match_status(match)
        metric_changes = _plan_metric_changes(
            parsed_row,
            event=batch.event,
            player=match.player,
        )
        has_conflict = any(
            change.get("action") == METRIC_ACTION_CONFLICT for change in metric_changes
        )
        row = AssessmentImportRow(
            batch=batch,
            row_key=parsed_row["row_key"],
            source_sheet=(parsed_row.get("source_rows") or [{}])[0].get("sheet", ""),
            source_row=(parsed_row.get("source_rows") or [{}])[0].get("row", 0) or 0,
            raw_identity=parsed_row["raw_identity"],
            player=match.player,
            roster_membership=match.roster_membership,
            match_status=match_status,
            validation_status=validation_status,
            conflict_status=(
                ASSESSMENT_CONFLICT_UNRESOLVED
                if has_conflict
                else ASSESSMENT_CONFLICT_NONE
            ),
            errors=row_errors,
            warnings=row_warnings,
            values_snapshot=parsed_row.get("values", []),
            metric_changes=metric_changes,
            raw_row={"source_rows": parsed_row.get("source_rows", [])},
            metadata={
                "match_reason": match.reason,
                "candidate_ids": [candidate.pk for candidate in match.candidates],
                "candidate_contexts": [
                    {
                        "player_id": context.player.pk,
                        "birth_year": context.birth_year,
                        "team": context.team,
                        "division": context.division,
                    }
                    for context in match.candidate_contexts
                ],
                "source_identifiers": parsed_row.get("source_identifiers", []),
            },
        )
        row.status = _legacy_row_status(row)
        row.action = _planned_row_action(row)
        row.save()
    if not batch.rows.filter(validation_status=ASSESSMENT_VALIDATION_VALID).exists():
        batch.validation_errors.append(
            _issue(
                "no_valid_player_rows",
                "No valid player rows are available for import.",
                blocking=True,
            )
        )
        batch.save(update_fields=["validation_errors", "updated_at"])
    batch.preview_snapshot = {
        **batch.preview_snapshot,
        "ranking_sheets": parsed.get("ranking_sheets", []),
    }
    return _refresh_batch_state(batch)


@transaction.atomic
def resolve_assessment_import_row(
    *,
    row: AssessmentImportRow,
    player: Player | None,
    skip: bool = False,
    refresh_batch: bool = True,
):
    """Resolve only identity state, or explicitly skip any invalid row."""
    row = (
        AssessmentImportRow.objects.select_for_update()
        .select_related("batch", "batch__event")
        .get(pk=row.pk)
    )
    if row.batch.status == ASSESSMENT_IMPORT_STATUS_COMMITTED:
        raise ValidationError("Committed assessment import rows cannot be changed.")
    if skip:
        row.player = None
        row.roster_membership = None
        row.status = ASSESSMENT_IMPORT_ROW_SKIPPED
        row.action = "skip"
    elif row.validation_status == ASSESSMENT_VALIDATION_INVALID:
        raise ValidationError(
            "Choosing a player cannot resolve data-validation errors; correct or skip the row."
        )
    elif player is None:
        raise ValidationError("Choose a player or skip the row.")
    else:
        row.player = player
        row.roster_membership = (
            player.roster_memberships.select_related("season_team")
            .filter(
                season_team__season=row.batch.event.season,
                is_active=True,
            )
            .order_by("-is_primary", "id")
            .first()
        )
        row.match_status = ASSESSMENT_MATCH_MATCHED
        parsed_row = {"values": row.values_snapshot}
        row.metric_changes = _plan_metric_changes(
            parsed_row,
            event=row.batch.event,
            player=player,
        )
        row.conflict_status = (
            ASSESSMENT_CONFLICT_UNRESOLVED
            if any(
                change.get("action") == METRIC_ACTION_CONFLICT
                for change in row.metric_changes
            )
            else ASSESSMENT_CONFLICT_NONE
        )
        row.status = _legacy_row_status(row)
        row.action = _planned_row_action(row)
    row.save()
    if refresh_batch:
        _refresh_batch_state(row.batch)
    return row


@transaction.atomic
def preserve_manual_override_conflicts(*, row: AssessmentImportRow, actor):
    """Resolve import conflicts by preserving every existing manual correction."""
    if not actor.is_staff and not actor.is_superuser:
        raise PermissionDenied("Only staff can resolve assessment import conflicts.")
    row = (
        AssessmentImportRow.objects.select_for_update()
        .select_related("batch")
        .get(pk=row.pk)
    )
    if row.batch.status == ASSESSMENT_IMPORT_STATUS_COMMITTED:
        raise ValidationError("Committed assessment import rows cannot be changed.")
    changes = deepcopy(row.metric_changes)
    conflict_found = False
    for change in changes:
        if change.get("action") == METRIC_ACTION_CONFLICT:
            conflict_found = True
            change["action"] = METRIC_ACTION_PROTECTED_MANUAL
            change["resolution"] = "preserve_manual"
    if not conflict_found:
        raise ValidationError("This row has no manual-override conflict to resolve.")
    row.metric_changes = changes
    row.conflict_status = ASSESSMENT_CONFLICT_RESOLVED
    row.status = _legacy_row_status(row)
    row.action = _planned_row_action(row)
    row.save()
    _refresh_batch_state(row.batch)
    return row


@transaction.atomic
def acknowledge_assessment_import_warnings(*, batch, actor, token):
    """Persist acknowledgement only for the current immutable preview state."""
    if not actor.is_staff and not actor.is_superuser:
        raise PermissionDenied("Only staff can acknowledge assessment warnings.")
    batch = AssessmentImportBatch.objects.select_for_update().get(pk=batch.pk)
    summary = _summary_counts(batch)
    if not summary.structurally_ready:
        raise ValidationError(
            "Resolve all blocking issues before acknowledging warnings."
        )
    if not batch.required_warning_codes:
        raise ValidationError("This assessment import has no warnings to acknowledge.")
    if not token or token != batch.acknowledgement_token:
        raise ValidationError(
            "Warning acknowledgement is stale; review the latest preview."
        )
    batch.warnings_acknowledged_at = timezone.now()
    batch.warnings_acknowledged_by = actor
    batch.metadata = {
        **batch.metadata,
        "acknowledged_token": batch.acknowledgement_token,
    }
    batch.save(
        update_fields=[
            "warnings_acknowledged_at",
            "warnings_acknowledged_by",
            "metadata",
            "updated_at",
        ]
    )
    return _summary_counts(batch)


def _metric_by_key(event) -> dict[str, AssessmentTemplateMetric]:
    return {
        template_metric.metric.key: template_metric
        for template_metric in event.template.template_metrics.select_related("metric")
    }


def _assign_snapshot_value(value: AssessmentValue, snapshot: dict):
    value.numeric_value = None
    value.rating_value = None
    value.text_value = ""
    value.choice_value = ""
    value.raw_value = snapshot.get("raw_value", "")
    value.normalized_value = snapshot.get("normalized_value", "")
    value.unit = snapshot.get("unit", "")
    value.source_sheet = snapshot.get("source_sheet", "")
    value.source_row = snapshot.get("source_row")
    value.source_column = snapshot.get("source_column", "")
    value.source_header = snapshot.get("source_header", snapshot.get("header", ""))
    value.source_kind = ASSESSMENT_VALUE_SOURCE_IMPORTED
    value.is_imported = True
    value.is_manual_override = False
    value.metadata = {
        "unit_status": snapshot.get("unit_status", "not_applicable"),
        "unit_source": snapshot.get("unit_source", ""),
        "zero_policy": snapshot.get("zero_policy", ZERO_ALLOW),
        "blank_policy": snapshot.get("blank_policy", BLANK_PRESERVE),
        "transformations": snapshot.get("transformations", []),
    }
    if snapshot.get("value_type") == ASSESSMENT_VALUE_TYPE_RATING:
        value.rating_value = Decimal(snapshot["normalized_value"])
        value.rating_scale_min = Decimal(str(snapshot["rating_scale_min"]))
        value.rating_scale_max = Decimal(str(snapshot["rating_scale_max"]))
    elif snapshot.get("value_type") == ASSESSMENT_VALUE_TYPE_NUMBER:
        value.numeric_value = Decimal(snapshot["normalized_value"])
        value.rating_scale_min = None
        value.rating_scale_max = None
    else:
        value.text_value = snapshot.get("text_value", snapshot.get("raw_value", ""))
        value.rating_scale_min = None
        value.rating_scale_max = None


def _apply_metric_change(
    *,
    player_assessment,
    template_metric,
    import_row,
    snapshot,
    change,
):
    existing = AssessmentValue.objects.filter(
        player_assessment=player_assessment,
        template_metric=template_metric,
    ).first()
    action = change["action"]
    if action in {
        METRIC_ACTION_SKIP,
        METRIC_ACTION_UNCHANGED,
        METRIC_ACTION_PROTECTED_MANUAL,
    }:
        return action
    if action == METRIC_ACTION_CLEAR:
        if existing:
            if existing.is_manual_override:
                raise ValidationError("Manual corrections cannot be cleared by import.")
            existing._allow_committed_change = True
            existing.delete()
        return action
    if action not in {METRIC_ACTION_CREATE, METRIC_ACTION_UPDATE}:
        raise ValidationError("Assessment import contains an unresolved metric action.")
    if existing and existing.is_manual_override:
        raise ValidationError("Assessment import cannot overwrite a manual correction.")
    value = existing or AssessmentValue(
        player_assessment=player_assessment,
        template_metric=template_metric,
    )
    _assign_snapshot_value(value, snapshot)
    value.import_row = import_row
    value._allow_committed_change = True
    value.save()
    return action


def _validate_commit_ready(batch: AssessmentImportBatch):
    if batch.status == ASSESSMENT_IMPORT_STATUS_COMMITTED:
        raise ValidationError("This assessment import has already been committed.")
    if batch.status != ASSESSMENT_IMPORT_STATUS_PREVIEWED:
        raise ValidationError("Only a successfully previewed import can be committed.")
    _validate_template_compatibility(batch.event, batch.import_template)
    if config_checksum(batch.config_snapshot) != batch.config_checksum:
        raise ValidationError("Frozen import configuration checksum is invalid.")
    summary = _summary_counts(batch)
    if not summary.can_commit:
        raise ValidationError(
            "Assessment import is not ready: resolve workbook, row, identity, conflict, and warning issues first."
        )
    if batch.validation_errors:
        raise ValidationError("Workbook-level validation errors block this import.")
    blocking_rows = batch.rows.exclude(status=ASSESSMENT_IMPORT_ROW_SKIPPED).filter(
        Q(validation_status=ASSESSMENT_VALIDATION_INVALID)
        | ~Q(match_status=ASSESSMENT_MATCH_MATCHED)
        | Q(conflict_status=ASSESSMENT_CONFLICT_UNRESOLVED)
    )
    if blocking_rows.exists():
        raise ValidationError(
            "Row-level validation or resolution issues block this import."
        )
    if not batch.rows.exclude(status=ASSESSMENT_IMPORT_ROW_SKIPPED).exists():
        raise ValidationError("At least one valid player row is required.")


@transaction.atomic
def commit_assessment_import_batch(
    *, batch: AssessmentImportBatch, actor
) -> AssessmentCommitResult:
    """Commit the frozen, fully resolved metric plan atomically."""
    if not actor.is_staff and not actor.is_superuser:
        raise PermissionDenied("Only staff can commit assessment imports.")
    batch = (
        AssessmentImportBatch.objects.select_for_update()
        .select_related(
            "event",
            "event__template",
            "event__scoring_profile",
            "import_template",
        )
        .get(pk=batch.pk)
    )
    _validate_commit_ready(batch)
    metrics = _metric_by_key(batch.event)
    created = updated = unchanged = skipped = 0
    value_counts = {
        METRIC_ACTION_CREATE: 0,
        METRIC_ACTION_UPDATE: 0,
        METRIC_ACTION_CLEAR: 0,
        METRIC_ACTION_UNCHANGED: 0,
        METRIC_ACTION_PROTECTED_MANUAL: 0,
    }
    for row in batch.rows.select_for_update().select_related("player"):
        if row.status == ASSESSMENT_IMPORT_ROW_SKIPPED:
            skipped += 1
            continue
        if not row.player_id:
            raise ValidationError("Every committed import row requires a player.")
        player_assessment = PlayerAssessment.objects.filter(
            player=row.player,
            event=batch.event,
        ).first()
        was_created = player_assessment is None
        if was_created:
            player_assessment = PlayerAssessment.objects.create(
                player=row.player,
                event=batch.event,
                roster_membership=row.roster_membership,
                import_batch=batch,
                source_row_key=row.row_key,
                status=ASSESSMENT_STATUS_DRAFT,
                metadata={"initial_import_batch_id": batch.pk},
            )
            created += 1
        snapshots = {
            snapshot["metric_key"]: snapshot for snapshot in row.values_snapshot
        }
        changed = False
        for change in row.metric_changes:
            if change.get("action") in {METRIC_ACTION_CONFLICT, METRIC_ACTION_INVALID}:
                raise ValidationError("Unresolved metric actions block this import.")
            template_metric = metrics.get(change.get("metric_key"))
            if template_metric is None:
                raise ValidationError(
                    f"Unknown assessment metric: {change.get('metric_key')}."
                )
            snapshot = snapshots.get(change["metric_key"])
            if snapshot is None:
                raise ValidationError("Frozen metric snapshot is missing.")
            action = _apply_metric_change(
                player_assessment=player_assessment,
                template_metric=template_metric,
                import_row=row,
                snapshot=snapshot,
                change=change,
            )
            if action in value_counts:
                value_counts[action] += 1
            if action in {
                METRIC_ACTION_CREATE,
                METRIC_ACTION_UPDATE,
                METRIC_ACTION_CLEAR,
            }:
                changed = True
        if was_created:
            player_assessment.status = ASSESSMENT_STATUS_COMMITTED
            player_assessment.save(update_fields=["status", "updated_at"])
        elif changed:
            updated += 1
        else:
            unchanged += 1
        row.status = ASSESSMENT_IMPORT_ROW_COMMITTED
        row.save(update_fields=["status", "updated_at"])
    batch.status = ASSESSMENT_IMPORT_STATUS_COMMITTED
    batch.committed_at = timezone.now()
    result = AssessmentCommitResult(
        processed=created + updated + unchanged + skipped,
        created=created,
        updated=updated,
        unchanged=unchanged,
        skipped=skipped,
        values_created=value_counts[METRIC_ACTION_CREATE],
        values_updated=value_counts[METRIC_ACTION_UPDATE],
        values_cleared=value_counts[METRIC_ACTION_CLEAR],
        values_unchanged=value_counts[METRIC_ACTION_UNCHANGED],
        values_protected=value_counts[METRIC_ACTION_PROTECTED_MANUAL],
    )
    batch.import_summary = asdict(result)
    batch.save(update_fields=["status", "committed_at", "import_summary", "updated_at"])
    for locked_object in [
        batch.event.template,
        batch.import_template,
        batch.event.scoring_profile,
    ]:
        if locked_object and not locked_object.is_locked:
            locked_object.is_locked = True
            locked_object.save(update_fields=["is_locked", "updated_at"])
    return result


def _correction_value_snapshot(value: AssessmentValue) -> dict:
    return _value_snapshot(value) or {}


@transaction.atomic
def correct_assessment_value(*, assessment_value, actor, reason, new_value):
    """Apply an audited staff correction without permitting import replacement."""
    if not actor.is_staff and not actor.is_superuser:
        raise PermissionDenied("Only staff can correct assessment values.")
    reason = str(reason or "").strip()
    if not reason:
        raise ValidationError("A correction reason is required.")
    value = (
        AssessmentValue.objects.select_for_update()
        .select_related("template_metric")
        .get(pk=assessment_value.pk)
    )
    previous = _correction_value_snapshot(value)
    metric = value.template_metric
    snapshot = {
        "value_type": metric.value_type,
        "raw_value": str(new_value),
        "normalized_value": str(new_value),
        "unit": metric.unit,
        "unit_status": metric.metadata.get("unit_status", "not_applicable"),
        "rating_scale_min": metric.rating_scale_min,
        "rating_scale_max": metric.rating_scale_max,
        "source_sheet": "",
        "source_row": None,
        "source_column": "",
        "source_header": metric.display_name,
        "zero_policy": metric.metadata.get("zero_policy", ZERO_ALLOW),
        "blank_policy": metric.metadata.get("blank_policy", BLANK_PRESERVE),
        "transformations": [],
    }
    if metric.value_type in {
        ASSESSMENT_VALUE_TYPE_NUMBER,
        ASSESSMENT_VALUE_TYPE_RATING,
    }:
        decimal_value = _decimal_or_none(new_value)
        if decimal_value is None:
            raise ValidationError("Correction must be a valid finite number.")
        if metric.min_value is not None and decimal_value < metric.min_value:
            raise ValidationError("Correction is below the metric minimum.")
        if metric.max_value is not None and decimal_value > metric.max_value:
            raise ValidationError("Correction is above the metric maximum.")
        if metric.value_type == ASSESSMENT_VALUE_TYPE_RATING:
            if decimal_value != decimal_value.to_integral_value():
                raise ValidationError("Rating corrections must be whole numbers.")
            if (
                metric.rating_scale_min is None
                or metric.rating_scale_max is None
                or decimal_value < metric.rating_scale_min
                or decimal_value > metric.rating_scale_max
            ):
                raise ValidationError("Correction is outside the rating scale.")
        snapshot["normalized_value"] = str(decimal_value)
    else:
        snapshot["text_value"] = str(new_value)
    _assign_snapshot_value(value, snapshot)
    value.source_kind = ASSESSMENT_VALUE_SOURCE_MANUAL_CORRECTED
    value.is_imported = False
    value.is_manual_override = True
    value.metadata = {
        **value.metadata,
        "correction_reason": reason,
        "corrected_at": timezone.now().isoformat(),
        "corrected_by_id": actor.pk,
    }
    value._allow_committed_change = True
    value.save()
    current = _correction_value_snapshot(value)
    AssessmentValueCorrection.objects.create(
        assessment_value=value,
        actor=actor,
        reason=reason,
        previous_snapshot=previous,
        new_snapshot=current,
        provenance={"source_kind": ASSESSMENT_VALUE_SOURCE_MANUAL_CORRECTED},
    )
    return value


def default_2026_13u_config() -> dict:
    """Return the versioned mapping verified against the supplied 2026 workbook."""
    return {
        "mapping_version": 1,
        "sheets": deepcopy(DEFAULT_2026_13U_DATA_SHEETS),
        "ranking_sheets": DEFAULT_2026_13U_RANKING_SHEETS,
        "limits": {
            "max_upload_bytes": DEFAULT_MAX_UPLOAD_BYTES,
            "max_archive_uncompressed_bytes": DEFAULT_MAX_UNCOMPRESSED_BYTES,
            "max_worksheets": DEFAULT_MAX_WORKSHEETS,
            "max_rows": DEFAULT_MAX_ROWS,
            "max_columns": DEFAULT_MAX_COLUMNS,
            "max_cell_text_length": DEFAULT_MAX_CELL_TEXT_LENGTH,
        },
        "notes": (
            "Ranking sheets are QA/provenance only. Physical measurement units are "
            "unverified because the workbook does not state them."
        ),
    }


def _expected_metric_rows(config):
    display_order = 0
    for sheet_config in config["sheets"]:
        for metric_config in sheet_config["metrics"]:
            display_order += 10
            yield sheet_config, metric_config, display_order


def _field_conflicts(instance, expected):
    conflicts = {}
    for field_name, expected_value in expected.items():
        actual_value = getattr(instance, field_name)
        if actual_value != expected_value:
            conflicts[field_name] = {
                "actual": actual_value,
                "expected": expected_value,
            }
    return conflicts


def _assert_fields(object_name, instance, expected):
    conflicts = _field_conflicts(instance, expected)
    if conflicts:
        raise ValidationError(
            f"Existing {object_name} configuration conflicts with the expected version: {conflicts}"
        )


def _configuration_plan(config) -> dict:
    metric_details = []
    for sheet_config, metric_config, _ in _expected_metric_rows(config):
        metric_details.append(
            {
                "key": metric_config["key"],
                "sheet": sheet_config["name"],
                "header": metric_config["header"],
                "value_type": metric_config["value_type"],
                "rating_scale": (
                    [
                        metric_config.get("rating_scale_min"),
                        metric_config.get("rating_scale_max"),
                    ]
                    if metric_config["value_type"] == ASSESSMENT_VALUE_TYPE_RATING
                    else None
                ),
                "unit": metric_config.get("unit", ""),
                "unit_status": metric_config.get("unit_status", "not_applicable"),
                "zero_policy": metric_config.get("zero_policy"),
                "blank_policy": metric_config.get("blank_policy"),
                "required_header": metric_config.get("required_header", True),
            }
        )
    return {
        "template": {
            "key": BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
            "version": 1,
        },
        "import_template": {
            "key": BOOTSTRAP_2026_13U_IMPORT_KEY,
            "version": 1,
            "config_checksum": config_checksum(config),
        },
        "scoring_profile": {
            "key": BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
            "version": 1,
        },
        "required_sheets": [
            sheet["name"] for sheet in config["sheets"] if sheet.get("required", True)
        ],
        "optional_sheets": [
            sheet["name"]
            for sheet in config["sheets"]
            if not sheet.get("required", True)
        ],
        "sheets": [
            {
                "name": sheet["name"],
                "required": sheet.get("required", True),
                "header_row": sheet.get("header_row", 1),
                "identity_column": sheet.get("identity_column", "Name"),
                "required_headers": [
                    metric["header"]
                    for metric in sheet.get("metrics", [])
                    if metric.get("required_header", True)
                ],
            }
            for sheet in config["sheets"]
        ],
        "metrics": metric_details,
    }


def _dry_run_state(*, label, instance, expected):
    conflicts = _field_conflicts(instance, expected) if instance else {}
    return {
        "object": label,
        "state": (
            "create" if instance is None else ("conflict" if conflicts else "present")
        ),
        "locked": bool(instance and getattr(instance, "is_locked", False)),
        "conflicts": conflicts,
    }


def _dry_run_configuration_states(config) -> list[dict]:
    template = AssessmentTemplate.objects.filter(
        key=BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
        version=1,
    ).first()
    states = [
        _dry_run_state(
            label="template",
            instance=template,
            expected={"name": "2026 VCB House 13U PeeWee Assessment"},
        )
    ]
    scoring_config = {"source": "workbook", "computed_scores": []}
    scoring_profile = AssessmentScoringProfile.objects.filter(
        key=BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
        version=1,
    ).first()
    states.append(
        _dry_run_state(
            label="scoring_profile",
            instance=scoring_profile,
            expected={
                "name": "2026 VCB House 13U PeeWee Assessment Scoring",
                "assessment_template_id": template.pk if template else None,
                "config": scoring_config,
            },
        )
    )
    import_template = AssessmentImportTemplate.objects.filter(
        key=BOOTSTRAP_2026_13U_IMPORT_KEY,
        version=1,
    ).first()
    states.append(
        _dry_run_state(
            label="import_template",
            instance=import_template,
            expected={
                "name": "2026 VCB House 13U PeeWee Assessment Workbook",
                "assessment_template_id": template.pk if template else None,
                "config": config,
            },
        )
    )
    for sheet_config, metric_config, display_order in _expected_metric_rows(config):
        metric = AssessmentMetricDefinition.objects.filter(
            key=metric_config["key"]
        ).first()
        value_type = metric_config["value_type"]
        states.append(
            _dry_run_state(
                label=f"metric:{metric_config['key']}",
                instance=metric,
                expected={
                    "name": metric_config["header"].strip(),
                    "default_value_type": value_type,
                    "default_unit": metric_config.get("unit", ""),
                    "metadata": {
                        "unit_status": metric_config.get(
                            "unit_status", "not_applicable"
                        ),
                        "unit_source": metric_config.get("unit_source", ""),
                    },
                },
            )
        )
        template_metric = None
        if template and metric:
            template_metric = AssessmentTemplateMetric.objects.filter(
                template=template,
                metric=metric,
            ).first()
        metric_metadata = {
            "source_sheet": sheet_config["name"],
            "source_header": metric_config["header"],
            "unit_status": metric_config.get("unit_status", "not_applicable"),
            "unit_source": metric_config.get("unit_source", ""),
            "zero_policy": metric_config.get("zero_policy", ZERO_ALLOW),
            "blank_policy": metric_config.get("blank_policy", BLANK_PRESERVE),
            "allowed_choices": metric_config.get("allowed_choices", []),
            "integer_only": metric_config.get("integer_only", False),
        }
        states.append(
            _dry_run_state(
                label=f"template_metric:{metric_config['key']}",
                instance=template_metric,
                expected={
                    "category": metric_config.get("category", sheet_config["name"]),
                    "display_name": metric_config["header"].strip(),
                    "display_order": display_order,
                    "value_type": value_type,
                    "unit": metric_config.get("unit", ""),
                    "direction": metric_config.get("direction", "neutral"),
                    "min_value": _decimal_or_none(metric_config.get("min_value")),
                    "max_value": _decimal_or_none(metric_config.get("max_value")),
                    "rating_scale_min": _decimal_or_none(
                        metric_config.get("rating_scale_min")
                    ),
                    "rating_scale_max": _decimal_or_none(
                        metric_config.get("rating_scale_max")
                    ),
                    "metadata": metric_metadata,
                },
            )
        )
    return states


@transaction.atomic
def ensure_2026_13u_assessment_configuration(*, dry_run: bool = False) -> dict:
    """Create exact versioned configuration or fail on any existing conflict."""
    config = default_2026_13u_config()
    plan = _configuration_plan(config)
    plan["states"] = _dry_run_configuration_states(config)
    if dry_run:
        return plan

    template, created = AssessmentTemplate.objects.get_or_create(
        key=BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
        version=1,
        defaults={"name": "2026 VCB House 13U PeeWee Assessment"},
    )
    if not created:
        _assert_fields(
            "assessment template",
            template,
            {"name": "2026 VCB House 13U PeeWee Assessment"},
        )
    scoring_config = {"source": "workbook", "computed_scores": []}
    scoring_profile, created = AssessmentScoringProfile.objects.get_or_create(
        key=BOOTSTRAP_2026_13U_ASSESSMENT_KEY,
        version=1,
        defaults={
            "name": "2026 VCB House 13U PeeWee Assessment Scoring",
            "assessment_template": template,
            "config": scoring_config,
        },
    )
    if not created:
        _assert_fields(
            "scoring profile",
            scoring_profile,
            {
                "name": "2026 VCB House 13U PeeWee Assessment Scoring",
                "assessment_template_id": template.pk,
                "config": scoring_config,
            },
        )
    import_template, created = AssessmentImportTemplate.objects.get_or_create(
        key=BOOTSTRAP_2026_13U_IMPORT_KEY,
        version=1,
        defaults={
            "name": "2026 VCB House 13U PeeWee Assessment Workbook",
            "assessment_template": template,
            "config": config,
        },
    )
    if not created:
        _assert_fields(
            "import template",
            import_template,
            {
                "name": "2026 VCB House 13U PeeWee Assessment Workbook",
                "assessment_template_id": template.pk,
                "config": config,
            },
        )

    for sheet_config, metric_config, display_order in _expected_metric_rows(config):
        value_type = metric_config["value_type"]
        metric, created = AssessmentMetricDefinition.objects.get_or_create(
            key=metric_config["key"],
            defaults={
                "name": metric_config["header"].strip(),
                "default_value_type": value_type,
                "default_unit": metric_config.get("unit", ""),
                "metadata": {
                    "unit_status": metric_config.get("unit_status", "not_applicable"),
                    "unit_source": metric_config.get("unit_source", ""),
                },
            },
        )
        if not created:
            _assert_fields(
                f"metric {metric_config['key']}",
                metric,
                {
                    "name": metric_config["header"].strip(),
                    "default_value_type": value_type,
                    "default_unit": metric_config.get("unit", ""),
                    "metadata": {
                        "unit_status": metric_config.get(
                            "unit_status", "not_applicable"
                        ),
                        "unit_source": metric_config.get("unit_source", ""),
                    },
                },
            )
        metric_metadata = {
            "source_sheet": sheet_config["name"],
            "source_header": metric_config["header"],
            "unit_status": metric_config.get("unit_status", "not_applicable"),
            "unit_source": metric_config.get("unit_source", ""),
            "zero_policy": metric_config.get("zero_policy", ZERO_ALLOW),
            "blank_policy": metric_config.get("blank_policy", BLANK_PRESERVE),
            "allowed_choices": metric_config.get("allowed_choices", []),
            "integer_only": metric_config.get("integer_only", False),
        }
        template_metric, created = AssessmentTemplateMetric.objects.get_or_create(
            template=template,
            metric=metric,
            defaults={
                "category": metric_config.get("category", sheet_config["name"]),
                "display_name": metric_config["header"].strip(),
                "display_order": display_order,
                "value_type": value_type,
                "unit": metric_config.get("unit", ""),
                "direction": metric_config.get("direction", "neutral"),
                "min_value": metric_config.get("min_value"),
                "max_value": metric_config.get("max_value"),
                "rating_scale_min": metric_config.get("rating_scale_min"),
                "rating_scale_max": metric_config.get("rating_scale_max"),
                "metadata": metric_metadata,
            },
        )
        if not created:
            _assert_fields(
                f"template metric {metric_config['key']}",
                template_metric,
                {
                    "category": metric_config.get("category", sheet_config["name"]),
                    "display_name": metric_config["header"].strip(),
                    "display_order": display_order,
                    "value_type": value_type,
                    "unit": metric_config.get("unit", ""),
                    "direction": metric_config.get("direction", "neutral"),
                    "min_value": _decimal_or_none(metric_config.get("min_value")),
                    "max_value": _decimal_or_none(metric_config.get("max_value")),
                    "rating_scale_min": _decimal_or_none(
                        metric_config.get("rating_scale_min")
                    ),
                    "rating_scale_max": _decimal_or_none(
                        metric_config.get("rating_scale_max")
                    ),
                    "metadata": metric_metadata,
                },
            )
    return plan


def assessment_records_for_player(player: Player):
    """Return staff-visible workbook assessment records for a player profile."""
    return (
        PlayerAssessment.objects.filter(player=player)
        .select_related("event", "event__season")
        .prefetch_related("values__template_metric")
        .order_by("-event__starts_on", "-created_at")
    )
