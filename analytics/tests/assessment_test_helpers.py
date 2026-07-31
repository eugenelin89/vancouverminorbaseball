from copy import deepcopy
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from analytics.models import (
    AssessmentEvent,
    AssessmentImportTemplate,
    AssessmentScoringProfile,
    AssessmentTemplate,
)
from analytics.services.assessment_import_service import (
    acknowledge_assessment_import_warnings,
    default_2026_13u_config,
    ensure_2026_13u_assessment_configuration,
)
from analytics.tests.helpers import Player, User, attach_player_to_season, create_season

ASSESSMENT_HEADERS = [
    "Name",
    "Home to 1st",
    "Broad Jump",
    "Lateral Jump",
    "Shotput",
    "Bat Speed",
    "Time 2 Contact",
    "Exit Velocity Avg.",
    "Exit Velocity Max",
    "Athletic Stance",
    "Balance Stride",
    "Barrel Level",
    "Launch Position",
    "Follow Through",
    "Readiness",
    "Footwork",
    "Glovework",
    "Athleticism",
    "Fundamental Throwing",
]

PITCHING_HEADERS = [
    "Name",
    "Velocity Avg.",
    "Velocity Max",
    "Pitch 1",
    "Pitch 2",
    "Pitch 3",
    "Pitch 4",
    "Athletic Movement",
    "Body Control",
    "Direction",
    "Repeatability",
    "Command2",
]


def assessment_row(name="Alex Example", **overrides):
    values = {
        "Home to 1st": 4.1,
        "Broad Jump": 82,
        "Lateral Jump": 60,
        "Shotput": 200,
        "Bat Speed": 55.2,
        "Time 2 Contact": 0.18,
        "Exit Velocity Avg.": 61.2,
        "Exit Velocity Max": 67.5,
        "Athletic Stance": 2,
        "Balance Stride": 2,
        "Barrel Level": 2,
        "Launch Position": 2,
        "Follow Through": 2,
        "Readiness": 2,
        "Footwork": 2,
        "Glovework": 2,
        "Athleticism": 2,
        "Fundamental Throwing": 2,
    }
    values.update(overrides)
    return [name, *[values[header] for header in ASSESSMENT_HEADERS[1:]]]


def pitching_row(name="Alex Example", **overrides):
    values = {
        "Velocity Avg.": 50,
        "Velocity Max": 53,
        "Pitch 1": "Fastball",
        "Pitch 2": "Changeup",
        "Pitch 3": "",
        "Pitch 4": "",
        "Athletic Movement": 2,
        "Body Control": 2,
        "Direction": 2,
        "Repeatability": 2,
        "Command2": 2,
    }
    values.update(overrides)
    return [name, *[values[header] for header in PITCHING_HEADERS[1:]]]


def workbook_bytes(
    *,
    assessment_rows=None,
    pitching_rows=None,
    include_assessment=True,
    include_pitching=True,
    assessment_headers=None,
    pitching_headers=None,
    assessment_header_row=2,
    pitching_header_row=2,
    extra_assessment_headers=None,
):
    workbook = Workbook()
    workbook.remove(workbook.active)
    if include_assessment:
        sheet = workbook.create_sheet("Assessment Data")
        for _ in range(1, assessment_header_row):
            sheet.append(["Athleticism Evaluation"])
        headers = list(assessment_headers or ASSESSMENT_HEADERS)
        headers.extend(extra_assessment_headers or [])
        sheet.append(headers)
        for row in assessment_rows or []:
            sheet.append([*row, *(["extra"] * len(extra_assessment_headers or []))])
    if include_pitching:
        sheet = workbook.create_sheet("Pitching Data ")
        for _ in range(1, pitching_header_row):
            sheet.append([])
        sheet.append(list(pitching_headers or PITCHING_HEADERS))
        for row in pitching_rows or []:
            sheet.append(row)
    if not workbook.worksheets:
        workbook.create_sheet("Empty")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def uploaded_workbook(content, name="assessment.xlsx"):
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def minimal_config(
    *,
    value_type="number",
    required_sheet=True,
    required_header=True,
    required_value=False,
    zero_policy="allow",
    blank_policy="preserve_existing",
    min_value=None,
    max_value=None,
    header="Metric",
    header_aliases=None,
):
    metric = {
        "header": header,
        "header_aliases": header_aliases or [],
        "key": "metric",
        "category": "Testing",
        "value_type": value_type,
        "required_header": required_header,
        "required_value": required_value,
        "zero_policy": zero_policy,
        "blank_policy": blank_policy,
        "unit": "",
        "unit_status": "not_applicable",
    }
    if value_type == "rating":
        metric.update(
            {
                "rating_scale_min": 1,
                "rating_scale_max": 3,
                "integer_only": True,
                "allowed_choices": [1, 2, 3],
            }
        )
    if min_value is not None:
        metric["min_value"] = min_value
    if max_value is not None:
        metric["max_value"] = max_value
    return {
        "mapping_version": 1,
        "sheets": [
            {
                "name": "Testing",
                "required": required_sheet,
                "header_row": 1,
                "identity_column": "Name",
                "max_rows": 20,
                "max_columns": 10,
                "metrics": [metric],
            }
        ],
        "limits": {
            "max_upload_bytes": 1024 * 1024,
            "max_worksheets": 5,
            "max_rows": 20,
            "max_columns": 10,
            "max_cell_text_length": 100,
        },
    }


def minimal_workbook(rows, *, header="Metric", extra_headers=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Testing"
    sheet.append(["Name", header, *(extra_headers or [])])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class AssessmentTestMixin:
    def setUp(self):
        super().setUp()
        self.staff = User.objects.create_user(
            username="assessment.staff", password="test", is_staff=True
        )
        self.user = User.objects.create_user(
            username="assessment.user", password="test"
        )
        self.season = create_season(name="Spring 2026", key="spring-2026")
        ensure_2026_13u_assessment_configuration()
        self.template = AssessmentTemplate.objects.get(key="2026-13u-house-assessment")
        self.import_template = AssessmentImportTemplate.objects.get(
            key="2026-13u-house-assessment-xlsx"
        )
        self.scoring_profile = AssessmentScoringProfile.objects.get(
            key="2026-13u-house-assessment"
        )
        self.event = AssessmentEvent.objects.create(
            name="Spring 2026 13U Assessment",
            season=self.season,
            division="13U House",
            template=self.template,
            scoring_profile=self.scoring_profile,
        )
        self.player = Player.objects.create(first_name="Alex", last_name="Example")
        self.membership = attach_player_to_season(
            self.player,
            self.season,
            team_name="Yankees",
            division="13U House",
        )

    def valid_upload(self, **assessment_overrides):
        return uploaded_workbook(
            workbook_bytes(
                assessment_rows=[assessment_row(**assessment_overrides)],
                pitching_rows=[pitching_row()],
            )
        )

    def acknowledge(self, batch):
        batch.refresh_from_db()
        if batch.required_warning_codes:
            acknowledge_assessment_import_warnings(
                batch=batch,
                actor=self.staff,
                token=batch.acknowledgement_token,
            )
        batch.refresh_from_db()
        return batch

    def custom_import_template(self, config=None, version=99):
        return AssessmentImportTemplate.objects.create(
            key="synthetic-assessment-import",
            name="Synthetic Assessment Import",
            version=version,
            assessment_template=self.template,
            config=deepcopy(config or default_2026_13u_config()),
        )
