from analytics.services.assessment_import_service import (
    create_assessment_import_batch,
    parse_assessment_workbook,
)
from analytics.services.assessment_matching_service import (
    MATCH_AMBIGUOUS,
    MATCH_EXACT_GLOBAL_ALIAS,
    MATCH_EXACT_GLOBAL_NAME,
    MATCH_EXACT_IDENTIFIER,
    MATCH_EXACT_ROSTER_ALIAS,
    MATCH_EXACT_ROSTER_NAME,
    MATCH_UNMATCHED,
    match_player_for_assessment,
)
from analytics.tests.assessment_test_helpers import (
    AssessmentTestMixin,
    assessment_row,
    minimal_config,
    minimal_workbook,
    pitching_row,
    uploaded_workbook,
    workbook_bytes,
)
from analytics.tests.helpers import (
    Player,
    TestCase,
    attach_player_to_season,
    create_season,
)
from players.models import PlayerAlias, PlayerSourceIdentifier


def row_issue_codes(parsed):
    return {issue["code"] for row in parsed["rows"] for issue in row.get("errors", [])}


class AssessmentMatchingTests(AssessmentTestMixin, TestCase):
    def test_matching_order_prefers_namespaced_identifier(self):
        PlayerSourceIdentifier.objects.create(
            player=self.player,
            source="registration",
            identifier_type="player_id",
            identifier_value="ABC-123",
        )
        result = match_player_for_assessment(
            raw_name="Different Name",
            event=self.event,
            source_identifiers=[
                {
                    "source": "registration",
                    "identifier_type": "player_id",
                    "identifier_value": "ABC-123",
                }
            ],
        )
        self.assertEqual(result.status, MATCH_EXACT_IDENTIFIER)
        self.assertEqual(result.player, self.player)

    def test_roster_name_and_alias_precede_global_matches(self):
        global_player = Player.objects.create(first_name="Alex", last_name="Example")
        roster_result = match_player_for_assessment(
            raw_name="Alex Example", event=self.event
        )
        self.assertEqual(roster_result.status, MATCH_EXACT_ROSTER_NAME)
        self.assertEqual(roster_result.player, self.player)

        PlayerAlias.objects.create(player=self.player, alias="A Example")
        PlayerAlias.objects.create(player=global_player, alias="A Example")
        alias_result = match_player_for_assessment(
            raw_name="A Example", event=self.event
        )
        self.assertEqual(alias_result.status, MATCH_EXACT_ROSTER_ALIAS)
        self.assertEqual(alias_result.player, self.player)

    def test_unique_global_name_and_alias_are_supported(self):
        outside = Player.objects.create(first_name="Outside", last_name="Player")
        result = match_player_for_assessment(
            raw_name="Outside Player", event=self.event
        )
        self.assertEqual(result.status, MATCH_EXACT_GLOBAL_NAME)
        self.assertEqual(result.player, outside)

        alias_player = Player.objects.create(first_name="Alias", last_name="Owner")
        PlayerAlias.objects.create(player=alias_player, alias="Unique Alias")
        result = match_player_for_assessment(raw_name="Unique Alias", event=self.event)
        self.assertEqual(result.status, MATCH_EXACT_GLOBAL_ALIAS)

    def test_duplicate_global_names_are_ambiguous_with_candidate_context(self):
        other_season = create_season(name="Fall 2025", key="fall-2025")
        first = Player.objects.create(
            first_name="Duplicate", last_name="Name", birth_year=2012
        )
        second = Player.objects.create(
            first_name="Duplicate", last_name="Name", birth_year=2013
        )
        attach_player_to_season(first, other_season, team_name="One", division="13U")
        attach_player_to_season(second, other_season, team_name="Two", division="13U")

        result = match_player_for_assessment(
            raw_name="Duplicate Name", event=self.event
        )

        self.assertEqual(result.status, MATCH_AMBIGUOUS)
        self.assertEqual(len(result.candidates), 2)
        self.assertEqual(
            {item.birth_year for item in result.candidate_contexts}, {2012, 2013}
        )

    def test_unmatched_never_creates_or_fuzzy_matches(self):
        player_count = Player.objects.count()
        result = match_player_for_assessment(raw_name="Alek Exampel", event=self.event)
        self.assertEqual(result.status, MATCH_UNMATCHED)
        self.assertEqual(Player.objects.count(), player_count)


class DuplicateWorkbookRowTests(AssessmentTestMixin, TestCase):
    def test_one_player_across_component_sheets_is_one_combined_row(self):
        parsed = parse_assessment_workbook(
            workbook_bytes(
                assessment_rows=[assessment_row()],
                pitching_rows=[pitching_row()],
            ),
            self.import_template.config,
        )
        self.assertEqual(len(parsed["rows"]), 1)
        self.assertEqual(len(parsed["rows"][0]["source_rows"]), 2)

    def test_duplicate_rows_within_each_sheet_are_blocking(self):
        assessment_duplicate = parse_assessment_workbook(
            workbook_bytes(
                assessment_rows=[assessment_row(), assessment_row()],
                pitching_rows=[],
            ),
            self.import_template.config,
        )
        self.assertIn(
            "duplicate_identity_in_sheet", row_issue_codes(assessment_duplicate)
        )

        pitching_duplicate = parse_assessment_workbook(
            workbook_bytes(
                assessment_rows=[assessment_row()],
                pitching_rows=[pitching_row(), pitching_row()],
            ),
            self.import_template.config,
        )
        self.assertIn(
            "duplicate_identity_in_sheet", row_issue_codes(pitching_duplicate)
        )

    def test_conflicting_duplicate_metric_values_are_reported(self):
        config = minimal_config()
        config["sheets"].append(
            {
                **config["sheets"][0],
                "name": "Testing Two",
            }
        )
        from io import BytesIO

        from openpyxl import Workbook

        workbook = Workbook()
        first = workbook.active
        first.title = "Testing"
        first.append(["Name", "Metric"])
        first.append(["Synthetic Player", 1])
        second = workbook.create_sheet("Testing Two")
        second.append(["Name", "Metric"])
        second.append(["Synthetic Player", 2])
        output = BytesIO()
        workbook.save(output)

        parsed = parse_assessment_workbook(output.getvalue(), config)
        self.assertIn("conflicting_duplicate_metric", row_issue_codes(parsed))

    def test_distinct_names_that_share_a_slug_are_not_silently_merged(self):
        parsed = parse_assessment_workbook(
            minimal_workbook([["Anne-Marie Test", 1], ["Anne Marie Test", 1]]),
            minimal_config(),
        )
        self.assertEqual(len(parsed["rows"]), 2)
        self.assertIn("identity_slug_collision", row_issue_codes(parsed))

    def test_duplicate_source_identifiers_are_blocked(self):
        config = minimal_config()
        config["sheets"][0]["source_identifiers"] = [
            {
                "header": "Registration ID",
                "source": "registration",
                "identifier_type": "player_id",
            }
        ]
        from io import BytesIO

        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Testing"
        sheet.append(["Name", "Metric", "Registration ID"])
        sheet.append(["One Player", 1, "same-id"])
        sheet.append(["Two Player", 2, "same-id"])
        output = BytesIO()
        workbook.save(output)

        parsed = parse_assessment_workbook(output.getvalue(), config)
        self.assertIn("duplicate_source_identifier", row_issue_codes(parsed))

    def test_preview_does_not_create_players_for_unmatched_rows(self):
        initial_count = Player.objects.count()
        batch = create_assessment_import_batch(
            file_obj=uploaded_workbook(
                workbook_bytes(
                    assessment_rows=[assessment_row(name="Unknown Person")],
                    pitching_rows=[],
                )
            ),
            event=self.event,
            import_template=self.import_template,
            uploaded_by=self.staff,
        )
        self.assertEqual(Player.objects.count(), initial_count)
        self.assertEqual(batch.rows.get().match_status, "unmatched")
